"""
build_sheet.py
Rebuilds __2026_BUDGETS_updated.xlsx from the original source with:
  - AI = HISTORY (inserted after AH)
  - AJ = <br+hist> scaffold (inserted after AI)
  - All scaffold formulas written explicitly — no insert_cols formula drift
  - AG renamed ACTION1, AG data cleaned (history migrated to AI)
  - Columns AK onward = original AI onward (shifted +2)
  - All formula references corrected for the shift
"""

import re, copy, sys
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cfi
from openpyxl.styles import PatternFill, Font, Alignment, Border

SRC = '/mnt/user-data/uploads/__2026_BUDGETS.xlsx'
DST = '/mnt/user-data/outputs/__2026_BUDGETS_updated.xlsx'

# ── Load source ───────────────────────────────────────────────────────────────

wb_src = load_workbook(SRC)
ws_src = wb_src.active

wb_raw = load_workbook(SRC, data_only=True)
ws_raw = wb_raw.active

def src_ci(l): return ws_src[l+'1'].column

# ── Column shift map ──────────────────────────────────────────────────────────
# Source cols A..AH (1..34) → same dst col
# Source cols AI..end (35..) → dst col + 2
# New cols: AI_DST=35 (HISTORY), AJ_DST=36 (<br+hist>)

AH_SRC = src_ci('AH')   # 34
AI_DST = AH_SRC + 1     # 35 - HISTORY
AJ_DST = AH_SRC + 2     # 36 - scaffold
HIST_L = gcl(AI_DST)    # 'AI'
SCAF_L = gcl(AJ_DST)    # 'AJ'

def dst_c(src_col):
    return src_col if src_col <= AH_SRC else src_col + 2

# Pre-compute destination letters for all referenced source columns
def dl(src_letter):
    return gcl(dst_c(cfi(src_letter)))

# ── Get AZ array formula template from source ────────────────────────────────

AZ_SRC_COL = src_ci('AZ')
AR_SRC_COL = src_ci('AR')
AU_SRC_COL = src_ci('AU')
AS_SRC_COL = src_ci('AS')

az_src_text = ws_src.cell(2, AZ_SRC_COL).value.text

# Tokenize AZ template: replace all row numbers with {r}, shift col refs > AH
def make_array_template(text, shift_after_src_col):
    # Replace row numbers with {r}
    text = re.sub(r'(?<=[A-Z])(\d+)', lambda m: '{r}' if int(m.group(1)) >= 2 else m.group(1), text)
    # Shift column refs
    def shift_col(m):
        col = m.group(1)
        suffix = m.group(2)
        idx = cfi(col)
        new_idx = idx + 2 if idx > shift_after_src_col else idx
        return gcl(new_idx) + suffix
    text = re.sub(r'\b([A-Z]{1,3})(\{r\}|\d)', shift_col, text)
    return text

AZ_TEMPLATE = make_array_template(az_src_text, AH_SRC)

AU_DST_L = dl('AU')
AS_DST_L = dl('AS')
AR_TEMPLATE = f'=_xlfn.IFS(NOT(ISBLANK({AU_DST_L}{{r}}))," (pgs. ",NOT(ISBLANK({AS_DST_L}{{r}}))," (pg. ", ISBLANK({AS_DST_L}{{r}}), "")'

# ── Style helpers ─────────────────────────────────────────────────────────────

def copy_style(c):
    return {k: copy.copy(getattr(c, k)) for k in ('font','fill','alignment','border')}
def apply_style(cell, style):
    for k,v in style.items(): setattr(cell, k, copy.copy(v))

data_style    = copy_style(ws_src['G1'])
scaffold_style = copy_style(ws_src['AD1'])

# ── Parse AG history ──────────────────────────────────────────────────────────

def parse_ag(val):
    if not val: return ('', '')
    parts = str(val).split('<br>')
    action1 = parts[0].strip()
    history = '<br>'.join(p.strip() for p in parts[1:] if p.strip())
    return (action1, history)

# ── Explicit scaffold formula map (per row r, using dst column letters) ───────

def scaffold(r):
    """All scaffold formulas keyed by dst column letter."""
    AK = dl('AI')   # old AI = <br> for italic note
    AL = dl('AJ')   # old AJ = <b><i> for italic note (DATA — skip)
    AM = dl('AK')   # old AK = italic note text (DATA — skip)
    AN = dl('AL')   # old AL = </i> scaffold
    AO = dl('AM')   # old AM = </b><br><br> scaffold  
    AP = dl('AN')   # old AN = <b> scaffold
    AQ = dl('AO')   # old AO = reduction plan label (DATA — skip)
    AR = dl('AP')   # old AP = reduction % (DATA — skip)
    AS_ = dl('AQ')  # old AQ = % scaffold
    AT = dl('AR')   # old AR = array formula — skip here
    AU_ = dl('AS')  # old AS = DFM page# (DATA — skip)
    AV = dl('AT')   # old AT = , scaffold
    AW_ = dl('AU')  # old AU = ## (DATA — skip)
    AX = dl('AV')   # old AV = ,<br> scaffold
    AY_ = dl('AW')  # old AW = (DATA — skip)
    AZ = dl('AX')   # old AX = )</b> scaffold
    BA_ = dl('AY')  # old AY = <br> scaffold
    BB = dl('AZ')   # old AZ = array formula — skip here
    BC_ = dl('BA')  # old BA = DFM text (DATA — skip)

    # Old AL formula referenced AK; now AK→AM (shifted), so AN references AM
    # Old AM formula referenced AO; now AO→AQ (shifted), so AO references AQ
    # etc.
    AK_data = dl('AK')  # italic note text col (was AK, now AM)
    AO_data = dl('AO')  # reduction label col (was AO, now AQ)
    AP_data = dl('AP')  # reduction % col (was AP, now AR)
    AS_data = dl('AS')  # DFM page# col (was AS, now AU)
    AU_data = dl('AU')  # ## col (was AU, now AW)
    AW_data = dl('AW')  # extra page ref col (was AW, now AY)
    BA_data = dl('BA')  # DFM text col (was BA, now BC)

    return {
        # Cols that didn't shift:
        'F':  f'=IF(LEFT(G{r},4)="FY26","<b>","<i>")',
        'H':  f'=IF(LEFT(G{r},4)="FY26","</b>","</i>")',
        'I':  '<i>',   # static data
        'K':  '</i>',  # static data
        'L':  f'=IF(ISBLANK(N{r}),"","<br>")',
        'M':  f'=IF(ISBLANK(R{r}),"<b>","<i>")',
        'O':  f'=IF(ISBLANK(R{r}),"</b>","</i>")',
        'P':  f'=IF(ISBLANK(R{r}),"","<br>")',
        'Q':  f'=IF(ISBLANK(R{r}),"","<b>")',
        'S':  f'=IF(ISBLANK(R{r}),"","</i>")',
        'T':  f'=IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"<b>","<i>"))',
        'U':  f'=IF(ISBLANK(V{r}),"","<a href=\'")',
        'W':  f"=IF(ISBLANK(V{r}),\"\",\"'>\")",
        'Y':  f'=IF(ISBLANK(X{r}),"","</a>")',
        'Z':  f'=IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"</b>","</i>"))',
        'AA': f'=IF(AB{r}="PASSED","<b>","")',
        'AC': f'=IF(AB{r}="PASSED","</b>","")',
        'AD': f'=IF(ISBLANK(AE{r}),"","<b>")',
        'AF': f'=IF(ISBLANK(AE{r}),"",": ")',
        'AH': f'=IF(ISBLANK(AE{r}),"","</b>")',
        # New inserted cols:
        SCAF_L: f'=IF(ISBLANK({HIST_L}{r}),"","<br>"&{HIST_L}{r})',
        # Shifted cols (original AI onward):
        AK:  f'=IF(ISBLANK(AE{r}),"","<br>")',              # old AI <br>
        # AL = old AJ = <b><i> — DATA, written from source
        # AM = old AK = italic note — DATA
        AN:  f'=IF(ISBLANK({AK_data}{r}),"","</i>")',       # old AL
        AO:  f'=IF(ISBLANK({AO_data}{r}),"","</b><br><br>")', # old AM
        AP:  f'=IF(ISBLANK({AO_data}{r}),"","<b>")',         # old AN
        # AQ = old AO = reduction label — DATA
        # AR = old AP = reduction % — DATA
        AS_: f'=IF(ISBLANK({AP_data}{r}),""," %")',          # old AQ (note: source has "%" not " %")
        # AT = old AR = array formula — handled separately
        # AU = old AS = DFM page# — DATA
        AV:  f'=IF(ISBLANK({AU_data}{r}),"",", ")',          # old AT
        # AW = old AU = ## — DATA
        AX:  f'=IF(ISBLANK({AW_data}{r}),"",",<br>")',       # old AV
        # AY = old AW — DATA
        AZ:  f'=IF(ISBLANK({AS_data}{r}),"",")</b>")',       # old AX
        BA_: f'=IF(ISBLANK({BA_data}{r}),"","<br>")',        # old AY
        # BB = old AZ = array formula — handled separately
        # BC = old BA = DFM text — DATA
        # E formula:
        'E':  f'=IF({AS_data}{r}="NA", _xlfn.TEXTJOIN("", TRUE, AD{r}:AH{r}), _xlfn.TEXTJOIN("", TRUE, AD{r}:{SCAF_L}{r}))',
        # A-D Flourish formulas (unchanged cols):
        'A':  f'=_xlfn.TEXTJOIN("", TRUE, F{r}:H{r})',
        'B':  f'=_xlfn.TEXTJOIN("", TRUE, I{r}:S{r})',
        'C':  f'=IF(ISBLANK(X{r}),"", _xlfn.TEXTJOIN("",TRUE,T{r}:Z{r}))',
        'D':  f'=IF(ISBLANK(AB{r}), "", _xlfn.TEXTJOIN("", TRUE, AA{r}:AC{r}))',
    }

# ── DATA columns (source col → dst col, value copied directly) ───────────────
DATA_COLS_SRC = {src_ci(l) for l in ['G','J','N','R','V','X','AB','AE','AG',
                                      'AK','AO','AP','AS','AU','AW','BA']}
# Note: AK/AO/AP/AS/AU/AW/BA are original source letters — their dst indices are shifted

# ── Build destination workbook ────────────────────────────────────────────────

from openpyxl import Workbook
wb_out = load_workbook(SRC)   # start from source to preserve styles/widths
ws_out = wb_out.active

# Insert 2 blank columns after AH in destination
ws_out.insert_cols(AH_SRC + 1, 2)

# Write HISTORY header + scaffold header
h1 = ws_out.cell(1, AI_DST); h1.value = 'HISTORY'; apply_style(h1, data_style)
ws_out.column_dimensions[HIST_L].width = 45
s1 = ws_out.cell(1, AJ_DST); s1.value = '<br+hist>'; apply_style(s1, scaffold_style)
ws_out.column_dimensions[SCAF_L].width = 5

# Rename AG
ws_out['AG1'].value = 'ACTION1'

# AR and AZ dst column indices
AR_DST_COL = dst_c(AR_SRC_COL)
AZ_DST_COL = dst_c(AZ_SRC_COL)

print("Writing rows...")
for r in range(2, ws_out.max_row + 1):
    sc = scaffold(r)

    # Write all explicit scaffold/Flourish formulas
    for col_letter, formula in sc.items():
        ws_out[f'{col_letter}{r}'].value = formula

    # Migrate AG → ACTION1 + HISTORY
    ag_raw = ws_raw.cell(r, src_ci('AG')).value
    action1, history = parse_ag(ag_raw)
    ws_out[f'AG{r}'].value = action1
    ws_out.cell(r, AI_DST).value = history if history else None

    # Fix array formulas (AR and AZ in dst, both shifted +2)
    ar_cell = ws_out.cell(r, AR_DST_COL)
    ar_cell.value = ArrayFormula(f'{gcl(AR_DST_COL)}{r}', AR_TEMPLATE.format(r=r))

    az_cell = ws_out.cell(r, AZ_DST_COL)
    az_cell.value = ArrayFormula(f'{gcl(AZ_DST_COL)}{r}', AZ_TEMPLATE.replace('{r}', str(r)))

print("Saving...")
wb_out.save(DST)
print(f"Saved: {DST}")

# ── Verify ────────────────────────────────────────────────────────────────────
wb_v = load_workbook(DST)
ws_v = wb_v.active

print("\n=== Full column map ===")
for c in range(1, ws_v.max_column + 1):
    h = ws_v.cell(1, c).value
    r2 = ws_v.cell(2, c).value
    if h is None and r2 is None: continue
    sample = str(r2)[:70] if r2 else 'empty'
    print(f"  {gcl(c):>3}  {repr(h):<20}  {sample}")
