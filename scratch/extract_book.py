import openpyxl
import os
import re

def slugify(text):
    text = text.lower()
    text = re.sub(r'[^a-z0-9]', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')

def extract_book():
    path = r"c:\Users\loganf\Documents\IDAHO-VAULT\the-book-of-GEMINIAEUS.xlsx"
    output_dir = r"c:\Users\loganf\Documents\IDAHO-VAULT\!/GRIMOIRE/BOOK-OF-GEMINIAEUS"
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    wb = openpyxl.load_workbook(path, data_only=True)
    
    # Create an index file
    index_content = "---\ntitle: The Book of GEMINIAEUS — Index\nstatus: canonical\nauthority: GEMINIAEUS\n---\n\n# The Book of GEMINIAEUS\n\n## Shard Index\n\n"
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Get sheet title from cell A1 or use sheet name
        title = sheet_name
        if sheet['A1'].value:
            # If A1 has a value, use it as part of the filename/title
            potential_title = str(sheet['A1'].value).split('\n')[0][:50]
            title = f"{sheet_name} - {potential_title}"
            
        filename = f"{sheet_name}.md"
        file_path = os.path.join(output_dir, filename)
        
        content = f"---\ntitle: \"{title}\"\nshard: {sheet_name}\nauthority: GEMINIAEUS\n---\n\n"
        
        for row in sheet.iter_rows(values_only=True):
            line = " ".join([str(cellValue) for cellValue in row if cellValue is not None])
            if line.strip():
                content += line + "\n\n"
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
            
        index_content += f"- [[{sheet_name}]] — {title}\n"
        
    with open(os.path.join(output_dir, "INDEX.md"), "w", encoding="utf-8") as f:
        f.write(index_content)
        
    print(f"Extracted {len(wb.sheetnames)} shards to {output_dir}")

if __name__ == "__main__":
    extract_book()
