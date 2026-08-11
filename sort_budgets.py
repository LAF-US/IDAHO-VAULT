"""
sort_budgets.py
Sort order:
  1. Action type: floor votes > committee actions > hearings > other
     - Floor votes: date desc first, then budget type
     - All others:  budget type first, then date desc
  2. Budget type: FY26 Supplemental > FY26 Reduction > FY27 Maintenance > FY27 Enhancement
  3. Date descending (most recent first); no-date rows sort last within their tier
  4. Sector (program) alphabetically
  5. Agency alphabetically
  6. Sub-agency alphabetically

Scaffold columns converted to formulas (never need manual editing):
  F  =IF(LEFT(G{r},4)="FY26","<b>","<i>")         — budget bold/italic
  H  =IF(LEFT(G{r},4)="FY26","</b>","</i>")
  L  =IF(ISBLANK(N{r}),"","<br>")                  — br between program and dept
  T  =IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"<b>","<i>"))  — bill bold/italic
  Z  =IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"</b>","</i>"))

AO stays as a manually editable data column (references a specific PDF).

Usage: python3 sort_budgets.py <input.xlsx> <output.xlsx>
       python3 sort_budgets.py  (sorts __2026_BUDGETS_updated.xlsx in place)
"""
import re, copy, sys
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

# ── Sort helpers ──────────────────────────────────────────────────────────────

def action_rank(desc, status):
    if not desc: desc = ''
    desc = str(desc).strip()
    status = str(status).strip() if status else ''
    if re.search(r'(House|Senate)\s+Vote', desc, re.I): return 0
    if 'Budget Setting' in desc: return 1
    if re.search(r'FY2\d\s+(reduction|maintenance)', desc, re.I): return 1
    if status == 'MOTION FAILED': return 1
    if 'Budget Hearing' in desc: return 2
    return 3

BUDGET_RANK = {
    'FY26 Supplemental': 0,
    'FY26 Reduction':    1,
    'FY27 Maintenance':  2,
    'FY27 Enhancement':  3,
}

def budget_rank(budget):
    return BUDGET_RANK.get(str(budget).strip(), 99)

def parse_date(d):
    """Return MMDD integer, or 0 if missing (sorts last when negated)."""
    if not d: return 0
    m = re.match(r'(\d{1,2})/(\d{1,2})', str(d).strip())
    return int(m.group(1)) * 100 + int(m.group(2)) if m else 0

# ── Scaffold formulas (columns previously requiring manual HTML tag entry) ─────

def scaffold_formulas(r):
    return {
        'F': f'=IF(LEFT(G{r},4)="FY26","<b>","<i>")',
        'H': f'=IF(LEFT(G{r},4)="FY26","</b>","</i>")',
        'L': f'=IF(ISBLANK(N{r}),"","<br>")',
        'T': f'=IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"<b>","<i>"))',
        'Z': f'=IF(ISBLANK(X{r}),"",IF(OR(AB{r}="PASSED",ISBLANK(N{r})),"</b>","</i>"))',
    }

# ── Array formula templates ───────────────────────────────────────────────────

AR_TEMPLATE = '=_xlfn.IFS(NOT(ISBLANK(AU{r}))," (pgs. ",NOT(ISBLANK(AS{r}))," (pg. ", ISBLANK(AS{r}), "")'

def get_az_template(ws, az_col):
    v = ws.cell(2, az_col).value
    text = v.text if isinstance(v, ArrayFormula) else str(v)
    text = re.sub(r'(?<=[A-Z])(\d+)', lambda m: '{r}' if int(m.group(1)) >= 2 else m.group(1), text)
    return text

# ── Main ──────────────────────────────────────────────────────────────────────

src = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs/__2026_BUDGETS_updated.xlsx'
dst = sys.argv[2] if len(sys.argv) > 2 else src

wb = load_workbook(src)
ws = wb.active
max_row = ws.max_row
max_col = ws.max_column

def ci(letter): return ws[letter + '1'].column
G=ci('G'); J=ci('J'); N=ci('N'); R=ci('R')
AB=ci('AB'); AE=ci('AE'); AG=ci('AG')

# Find array formula columns by header (letters may shift as columns are added)
AR_COL = next(c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == '<>(_')
AZ_COL = next(c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == 'TEXT')

AZ_TEMPLATE = get_az_template(ws, AZ_COL)

wb_d = load_workbook(src, data_only=True)
ws_d = wb_d.active

def snapshot(ws, r):
    cells = []
    for c in range(1, max_col + 1):
        cell = ws.cell(r, c)
        v = cell.value
        is_array = isinstance(v, ArrayFormula)
        cells.append({
            'value': None if is_array else v,
            'is_array': is_array,
            'col': c,
            'col_letter': cell.column_letter,
            'number_format': cell.number_format,
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'alignment': copy.copy(cell.alignment),
            'border': copy.copy(cell.border),
        })
    return cells

# Build sort keys
sort_keys = []
for r in range(2, max_row + 1):
    action = action_rank(ws_d.cell(r, AG).value, ws_d.cell(r, AB).value)
    date   = -parse_date(ws_d.cell(r, AE).value)   # negate for descending; 0 → no date, sorts last
    budget = budget_rank(ws_d.cell(r, G).value)
    sector = str(ws_d.cell(r, J).value or '')
    agency = str(ws_d.cell(r, N).value or '')
    subag  = str(ws_d.cell(r, R).value or '')

    # Floor votes: date first, then budget type
    # All others: budget type first, then date
    if action == 0:
        sort_keys.append((action, date, budget, sector, agency, subag))
    else:
        sort_keys.append((action, budget, date, sector, agency, subag))

print('Snapshotting...')
snapshots = [snapshot(ws, r) for r in range(2, max_row + 1)]

combined = sorted(zip(sort_keys, snapshots), key=lambda x: x[0])
sorted_snaps = [s for _, s in combined]

print('Writing...')
for new_idx, cells in enumerate(sorted_snaps):
    er = new_idx + 2
    scaffolds = scaffold_formulas(er)

    for props in cells:
        c   = props['col']
        cl  = props['col_letter']
        cell = ws.cell(er, c)

        if props['is_array']:
            text = AR_TEMPLATE.format(r=er) if c == AR_COL else AZ_TEMPLATE.replace('{r}', str(er))
            cell.value = ArrayFormula(f'{cl}{er}', text)

        elif cl in scaffolds:
            cell.value = scaffolds[cl]

        else:
            v = props['value']
            if isinstance(v, str) and v.startswith('='):
                v = re.sub(
                    r'([A-Z]+)(\d+)',
                    lambda m: m.group(1) + str(er) if int(m.group(2)) >= 2 else m.group(0),
                    v
                )
            cell.value = v

        cell.number_format = props['number_format']
        cell.font          = props['font']
        cell.fill          = props['fill']
        cell.alignment     = props['alignment']
        cell.border        = props['border']

wb.save(dst)
print(f'Saved: {dst}')

# Verify
wb2 = load_workbook(dst, data_only=True)
ws2 = wb2.active
print('\n=== Sort check: first 20 rows ===')
print(f'{"Row":>4}  {"Action":<20} {"Budget":<22} {"Date":<7} {"Agency"}')
print('-' * 90)
import re as _re
def _action_label(desc, status):
    desc = str(desc or '').strip(); status = str(status or '').strip()
    if _re.search(r'(House|Senate)\s+Vote', desc, _re.I): return 'Floor Vote'
    if 'Budget Setting' in desc or _re.search(r'FY2\d\s+(reduction|maintenance)', desc, _re.I): return 'Committee Action'
    if status == 'MOTION FAILED': return 'Committee Action'
    if 'Budget Hearing' in desc: return 'Committee Hearing'
    return 'Other'
for r in range(2, 22):
    action = _action_label(ws2.cell(r, AG).value, ws2.cell(r, AB).value)
    date   = ws2.cell(r, AE).value or ''
    budget = ws2.cell(r, G).value or ''
    agency = ws2.cell(r, N).value or ''
    print(f'{r:>4}  {action:<20} {str(budget):<22} {str(date):<7} {agency}')
