import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


BILL_ID_RE = re.compile(r"([HS][0-9]{3,4}[a-z]?)", re.IGNORECASE)
ADVANCED_TOKENS = (
    "PASSED*",
    "SIGNED",
    "3RD RDG",
    "2ND RDG",
    "ENGROS",
    "AMND",
    "AMEN",
    "ENROLL",
    "TRANSP",
    "RPT OUT",
    "RLS",
    "RET TO",
    "TO 14TH",
    "TO 10TH",
    "TO 3RD",
)


def simplify_status(raw_status: str) -> str:
    """Strip branch (H/S), dates (MM/DD), and vote tallies (A-B-C) from status."""
    if not raw_status:
        return ""
    # Remove branch indicator: (H) or (S)
    s = re.sub(r"^\([HS]\)\s*", "", raw_status.strip(), flags=re.IGNORECASE)
    # Remove date fragments: MM/DD or M/D
    s = re.sub(r"\b\d{1,2}/\d{1,2}\b", "", s)
    # Remove vote counts: A-B-C or [A-B-C]
    s = re.sub(r"\[?\b\d{1,3}-\d{1,3}-\d{1,3}\b\]?", "", s)
    # Collapse whitespace
    return re.sub(r"\s+", " ", s).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Sync the !_2026_BUDGETS.xlsx event columns from minidata. "
            "This script shifts historical events to the right (AE/AG -> AH/AI, etc.) "
            "and sets the STATUS_OVERRIDE (BB) when necessary."
        )
    )
    parser.add_argument(
        "--minidata",
        type=Path,
        help="Explicit path to a minidata CSV. Defaults to the freshest local minidata*.csv.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("!_2026_BUDGETS.xlsx"),
        help="Path to the budget tracker workbook.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report the changes without writing the workbook.",
    )
    parser.add_argument(
        "--snapshot-date",
        help=(
            "Optional snapshot date for review output (YYYY-MM-DD or MM/DD). "
            "Used only for manual event-history warnings."
        ),
    )
    return parser.parse_args()


def find_latest_minidata(explicit: Path | None) -> Path:
    if explicit:
        return explicit

    candidates = []
    search_roots = [Path.cwd(), Path.home() / "Downloads"]
    for root in search_roots:
        if not root.exists():
            continue
        candidates.extend(root.glob("minidata*.csv"))

    if not candidates:
        raise FileNotFoundError("No local minidata*.csv file found.")

    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_minidata_rows(path: Path) -> dict[str, tuple[str, str, str]]:
    rows: dict[str, tuple[str, str, str]] = {}
    with path.open("r", encoding="cp1252", newline="") as handle:
        for row in csv.reader(handle):
            if not row or len(row) < 3:
                continue
            bill_id = row[0].strip().upper()
            if not bill_id:
                continue
            title = row[1].strip()
            status = row[2].strip()
            vote = row[3].strip() if len(row) > 3 else ""
            rows[bill_id] = (title, status, vote)
    return rows


def map_tracker_status(raw_status: str) -> str:
    normalized = (raw_status or "").strip().upper()
    if not normalized:
        return ""
    if "FAILED" in normalized:
        return "FAILED"
    if normalized == "LAW+":
        return "PASSED"
    if normalized.startswith("TO GOV"):
        return "ADVANCED"
    if any(token in normalized for token in ADVANCED_TOKENS):
        return "ADVANCED"
    return "INTRODUCED"


def extract_bill_id(url_value: object, label_value: object) -> str | None:
    for candidate in (str(url_value or ""), str(label_value or "")):
        match = BILL_ID_RE.search(candidate)
        if match:
            return match.group(1).upper()
    return None


def normalize_snapshot_date(snapshot_date: str | None) -> str | None:
    if not snapshot_date:
        return None
    if re.fullmatch(r"\d{2}/\d{2}", snapshot_date):
        return snapshot_date
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", snapshot_date):
        return snapshot_date[5:7] + "/" + snapshot_date[8:10]
    return snapshot_date


def main() -> int:
    args = parse_args()
    minidata_path = find_latest_minidata(args.minidata)
    workbook_path = args.workbook

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    minidata_rows = load_minidata_rows(minidata_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook[workbook.sheetnames[0]]
    snapshot_date = normalize_snapshot_date(args.snapshot_date)

    changes: list[tuple[int, str, str, str, str, str]] = []
    for row_number in range(2, sheet.max_row + 1):
        bill_id = extract_bill_id(sheet[f"V{row_number}"].value, sheet[f"X{row_number}"].value)
        if not bill_id or bill_id not in minidata_rows:
            continue

        _title, raw_status, vote = minidata_rows[bill_id]
        
        current_date_val = sheet[f"AE{row_number}"].value
        current_action_val = sheet[f"AG{row_number}"].value
        current_date = (str(current_date_val) if current_date_val else "").strip()
        current_summary = (str(current_action_val) if current_action_val else "").strip()
        
        # Determine if we should record a change
        # If the minidata snapshot doesn't offer a new date, but the raw_status changed:
        if snapshot_date:
            if current_date == snapshot_date and current_summary == raw_status:
                continue
        else:
            if current_summary == raw_status:
                continue

        simple_status = simplify_status(raw_status)
        changes.append((row_number, bill_id, current_summary, raw_status, raw_status, simple_status))
        
        if not args.dry_run:
            # Shift events rightward
            sheet[f"AT{row_number}"] = sheet[f"AR{row_number}"].value
            sheet[f"AU{row_number}"] = sheet[f"AS{row_number}"].value
            
            sheet[f"AR{row_number}"] = sheet[f"AP{row_number}"].value
            sheet[f"AS{row_number}"] = sheet[f"AQ{row_number}"].value
            
            sheet[f"AP{row_number}"] = sheet[f"AN{row_number}"].value
            sheet[f"AQ{row_number}"] = sheet[f"AO{row_number}"].value
            
            sheet[f"AN{row_number}"] = sheet[f"AL{row_number}"].value
            sheet[f"AO{row_number}"] = sheet[f"AM{row_number}"].value
            
            sheet[f"AL{row_number}"] = sheet[f"AJ{row_number}"].value
            sheet[f"AM{row_number}"] = sheet[f"AK{row_number}"].value
            
            sheet[f"AJ{row_number}"] = sheet[f"AH{row_number}"].value
            sheet[f"AK{row_number}"] = sheet[f"AI{row_number}"].value
            
            sheet[f"AH{row_number}"] = sheet[f"AE{row_number}"].value
            sheet[f"AI{row_number}"] = sheet[f"AG{row_number}"].value
            
            # Write new event
            if snapshot_date:
                sheet[f"AE{row_number}"] = snapshot_date
            sheet[f"AG{row_number}"] = raw_status
            
            # Write STATUS_OVERRIDE if applicable
            # AB represents standard statuses, BB is strictly the override slot.
            if "WITHDRAWN" in raw_status.upper() or "MOTION FAILED" in raw_status.upper():
                sheet[f"BB{row_number}"] = raw_status

            # Write simplified status to Column E (5th column)
            sheet[f"E{row_number}"] = simple_status

    if not args.dry_run and changes:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(workbook_path)

    print(f"minidata: {minidata_path}")
    print(f"workbook: {workbook_path}")
    print(f"row_changes: {len(changes)}")

    seen_bills: set[str] = set()
    for row_number, bill_id, old_action, new_action, raw_status, simple_status in changes:
        if bill_id in seen_bills:
            continue
        seen_bills.add(bill_id)
        print(
            f"bill {bill_id}: {old_action} -> {new_action} "
            f"[Simple: {simple_status}] "
            f"(first row {row_number})"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
