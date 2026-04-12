#!/usr/bin/env python3
"""
Bind AI chat exports into local-only xlsx workbooks.

Supported providers:
- anthropic / claude: expects a Claude export zip or conversations.json
- openai / chatgpt: expects a ChatGPT export zip or conversations.json

Output:
- Sheet1: index / census
- one worksheet per conversation in source order

This script is intentionally local-first. The resulting xlsx files are
expected to remain gitignored and private.
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook


@dataclass
class Message:
    role: str
    text: str
    created_at: str = ""


@dataclass
class Conversation:
    title: str
    created_at: str
    source_id: str
    messages: list[Message]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Bind Anthropic/OpenAI exports into xlsx workbooks."
    )
    parser.add_argument(
        "--provider",
        required=True,
        choices=["anthropic", "claude", "openai", "chatgpt"],
        help="Export provider to parse.",
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to an export zip or conversations.json file.",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Path to the output xlsx workbook.",
    )
    return parser.parse_args()


def load_export_json(input_path: Path) -> Any:
    if input_path.suffix.lower() == ".json":
        return json.loads(input_path.read_text(encoding="utf-8-sig"))

    if input_path.suffix.lower() != ".zip":
        raise ValueError(f"Unsupported input type: {input_path.suffix}")

    with zipfile.ZipFile(input_path) as archive:
        match = next(
            (name for name in archive.namelist() if name.endswith("conversations.json")),
            None,
        )
        if not match:
            raise FileNotFoundError("Could not find conversations.json inside zip export")
        with archive.open(match) as handle:
            payload = handle.read().decode("utf-8-sig")
    return json.loads(payload)


def safe_sheet_title(raw: str, fallback: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]\*\:/\\\?]", "_", (raw or "").strip())[:31] or fallback
    candidate = title
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def stringify_timestamp(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(value, UTC).isoformat()
        except Exception:
            return str(value)
    return str(value)


def coerce_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, str):
                parts.append(item)
            elif isinstance(item, dict):
                item_type = item.get("type")
                if item_type == "text":
                    text = item.get("text")
                    if isinstance(text, str):
                        parts.append(text)
                elif isinstance(item.get("text"), str):
                    parts.append(item["text"])
        return "\n".join(part for part in parts if part).strip()
    if isinstance(value, dict):
        if isinstance(value.get("text"), str):
            return value["text"].strip()
        if isinstance(value.get("parts"), list):
            return coerce_text(value["parts"])
    return str(value).strip()


def parse_anthropic_export(data: Any) -> list[Conversation]:
    conversations = data if isinstance(data, list) else data.get("conversations", [])
    result: list[Conversation] = []

    for index, conv in enumerate(conversations, start=1):
        title = conv.get("name") or conv.get("title") or f"Conversation {index}"
        created_at = stringify_timestamp(conv.get("created_at") or conv.get("createdAt"))
        source_id = str(conv.get("uuid") or conv.get("id") or index)
        raw_messages = conv.get("chat_messages") or conv.get("messages") or []
        messages: list[Message] = []

        for raw_msg in raw_messages:
            role = str(raw_msg.get("sender") or raw_msg.get("role") or "unknown")
            text = coerce_text(raw_msg.get("text") or raw_msg.get("content"))
            if not text:
                continue
            msg_created = stringify_timestamp(
                raw_msg.get("created_at") or raw_msg.get("createdAt")
            )
            messages.append(Message(role=role, text=text, created_at=msg_created))

        result.append(
            Conversation(
                title=title,
                created_at=created_at,
                source_id=source_id,
                messages=messages,
            )
        )

    return result


def extract_openai_message_text(message: dict[str, Any]) -> str:
    content = message.get("content")
    if not isinstance(content, dict):
        return ""
    parts = content.get("parts")
    return coerce_text(parts)


def parse_openai_export(data: Any) -> list[Conversation]:
    conversations = data if isinstance(data, list) else data.get("conversations", [])
    result: list[Conversation] = []

    for index, conv in enumerate(conversations, start=1):
        title = conv.get("title") or f"Conversation {index}"
        created_at = stringify_timestamp(conv.get("create_time"))
        source_id = str(conv.get("id") or index)
        mapping = conv.get("mapping") or {}
        rows: list[tuple[float, Message]] = []

        for node in mapping.values():
            if not isinstance(node, dict):
                continue
            message = node.get("message")
            if not isinstance(message, dict):
                continue
            author = message.get("author") or {}
            role = str(author.get("role") or "unknown")
            text = extract_openai_message_text(message)
            if not text:
                continue
            msg_time = message.get("create_time") or node.get("create_time") or 0
            rows.append(
                (
                    float(msg_time or 0),
                    Message(
                        role=role,
                        text=text,
                        created_at=stringify_timestamp(msg_time),
                    ),
                )
            )

        rows.sort(key=lambda item: item[0])
        result.append(
            Conversation(
                title=title,
                created_at=created_at,
                source_id=source_id,
                messages=[message for _, message in rows],
            )
        )

    return result


def build_workbook(
    provider: str, conversations: list[Conversation], input_path: Path, output_path: Path
) -> None:
    workbook = Workbook()
    index = workbook.active
    index.title = "Sheet1"
    index.append([provider.capitalize()])
    index.append([f"Bound on {datetime.now(UTC).isoformat()}"])
    index.append(["title", "created_at", "messages", "source_id"])

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["provider", provider])
    metadata.append(["source_path", str(input_path)])
    metadata.append(["conversation_count", len(conversations)])
    metadata.append(["message_count", sum(len(conv.messages) for conv in conversations)])
    metadata.append(["bound_at", datetime.now(UTC).isoformat()])

    used_titles = {"Sheet1", "Metadata"}
    for conv in conversations:
        index.append([conv.title, conv.created_at, len(conv.messages), conv.source_id])
        sheet = workbook.create_sheet(
            safe_sheet_title(conv.title, f"Conversation_{len(workbook.sheetnames)}", used_titles)
        )
        sheet.append([provider.capitalize()])
        sheet.append([conv.title])
        if conv.created_at:
            sheet.append([f"Created: {conv.created_at}"])
        sheet.append([])
        for message in conv.messages:
            header = message.role.title()
            if message.created_at:
                header = f"{header} ({message.created_at})"
            sheet.append([header])
            for line in (message.text.splitlines() or [""]):
                sheet.append([line])
            sheet.append([])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    provider = args.provider.lower()
    data = load_export_json(args.input)

    if provider in {"anthropic", "claude"}:
        conversations = parse_anthropic_export(data)
        provider_name = "Claudius"
    else:
        conversations = parse_openai_export(data)
        provider_name = "Codices"

    build_workbook(provider_name, conversations, args.input, args.output)
    print(f"bound: {args.output}")
    print(f"conversations: {len(conversations)}")


if __name__ == "__main__":
    main()
